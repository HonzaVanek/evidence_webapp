from django.core.mail import send_mail, EmailMessage, EmailMultiAlternatives, get_connection
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.shortcuts import render, get_object_or_404, redirect
from .models import Pojistenec, Pojisteni, TypPojisteni, Contact, EmailTemplate, EmailCampaign, EmailDelivery, EmailImage, ContactGroup
from .forms import PojistenecForm, PojisteniForm, TypPojisteniForm, BulkUploadForm, RegistraceForm, ContactForm, ContactImportForm, EmailTemplateForm, SendCampaignForm, EmailImageUploadForm, ContactGroupForm
from django.http import HttpResponseRedirect, HttpResponse, HttpResponseForbidden
from django.contrib import messages
from django.urls import reverse
from django.utils import timezone
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.tokens import default_token_generator
from django.contrib.sites.shortcuts import get_current_site
from django.contrib.auth.views import LoginView
from .forms import VlastniLoginForm, RegistraceForm
from django.template import Context, Template
from django.template.loader import render_to_string
from django.conf import settings
from openpyxl import load_workbook, Workbook
import requests
from django.db import transaction, IntegrityError
from django.db.models import Q, Sum
from django import forms
from rest_framework import viewsets
from .serializers import PojistenecSerializer
from rest_framework.permissions import AllowAny, IsAdminUser
import qrcode
from PIL import Image
import os
import matplotlib
matplotlib.use("Agg") 
import matplotlib.pyplot as plt
from matplotlib.patches import PathPatch
from matplotlib.patches import Wedge
from matplotlib.path import Path
import csv
from io import TextIOWrapper


# Create your views here.

class PojistenecViewSet(viewsets.ModelViewSet):
    queryset = Pojistenec.objects.all()
    serializer_class = PojistenecSerializer
    permission_classes = [AllowAny]

    def get_permissions(self):
        if self.request.method in ['POST', 'PUT', 'PATCH', 'DELETE']:
            return [IsAdminUser()]
        return super().get_permissions()

def index(request):
    return render(request, 'pojistenci/index.html')

@login_required
def pojistenci_vypis(request):
    search_query = request.GET.get('search', '').strip()
    if search_query:
        pojistenci = Pojistenec.objects.filter(
            Q(first_name__icontains=search_query) | 
            Q(last_name__icontains=search_query) |
            Q(address_street__icontains=search_query) |
            Q(address_city__icontains=search_query) |
            Q(psc__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query)
        ).order_by('last_name', 'first_name')
    else:
        pojistenci = Pojistenec.objects.all().order_by('last_name', 'first_name')
    
    paginator = Paginator(pojistenci, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'pojistenci/pojistenci.html', {'pojistenci': pojistenci, 'page_obj': page_obj, 'total_count': paginator.count, 'search_query': search_query})

@login_required
def pojistenec_detail(request, id):
    pojistenec = get_object_or_404(Pojistenec, id=id)
    
    return render(request, 'pojistenci/pojistenec_detail.html', {
        'jmeno': pojistenec.first_name,
        'prijmeni': pojistenec.last_name,
        'ulice' : pojistenec.address_street,
        'mesto' : pojistenec.address_city,
        'psc': pojistenec.psc,
        'tel': pojistenec.phone,
        'email': pojistenec.email,
        'id': pojistenec.id,
        'foto': pojistenec.foto,
        'pojisteni' : pojistenec.pojisteni.all()
        })

@login_required
def novy_pojistenec(request):
    if request.method == 'POST':
        form = PojistenecForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pojištěnec byl úspěšně přidán.')
            return redirect('pojistenci')
        
    else:
        form = PojistenecForm()

    return render(request, 'pojistenci/novy_pojistenec.html', {'form': form})

@login_required
def smaz_pojistence(request, pojistenec_id):
    pojistenec = get_object_or_404(Pojistenec, id = pojistenec_id)
    pojistenec.delete()
    messages.success(request, 'Pojištěnec byl úspěšně odstraněn.', extra_tags='deleted')
    return HttpResponseRedirect ('/pojistenci')

@login_required
def uprav_pojistence(request, pojistenec_id):
    pojistenec = get_object_or_404(Pojistenec, id = pojistenec_id)
    if request.method == 'POST':
        pojistenec.first_name = request.POST['pojistenec.first_name']
        pojistenec.last_name = request.POST['pojistenec.last_name']
        pojistenec.address_street = request.POST['pojistenec.address_street']
        pojistenec.address_city = request.POST['pojistenec.address_city']
        pojistenec.psc = request.POST['pojistenec.psc']
        pojistenec.phone = request.POST['pojistenec.phone']
        pojistenec.email = request.POST['pojistenec.email']
        if 'foto' in request.FILES:
            pojistenec.foto = request.FILES['foto']
        pojistenec.save()
        return redirect ('pojistenec_detail', id=pojistenec_id)
    
    return render(request, 'pojistenci/uprav-pojistence.html', {
        'pojistenec': pojistenec
    })

@login_required
def pridat_pojisteni(request, pojistenec_id):
    pojistenec = get_object_or_404(Pojistenec, id = pojistenec_id)
    if request.method == 'POST':
        form = PojisteniForm(request.POST)
        if form.is_valid():
            pojisteni = form.save(commit=False)
            pojisteni.pojistenec = pojistenec
            pojisteni.save()
            messages.success(request, 'Pojištění bylo přidáno.')
            return redirect ('pojistenec_detail', id=pojistenec_id)
    
    else:
        form = PojisteniForm()
    
    return render(request, 'pojistenci/pridat_pojisteni.html', {'form' : form, 'pojistenec' : pojistenec})

@login_required
def pojisteni_vypis(request):
    pojisteni = Pojisteni.objects.all().order_by('platnost_do')
    paginator = Paginator(pojisteni, 10)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    today = timezone.now().date()

    return render(request, 'pojistenci/pojisteni.html', {'pojisteni': pojisteni, 'page_obj': page_obj, 'total_count': paginator.count, 'today': today})

@staff_member_required(login_url='login')
def smaz_typ_pojisteni(request, typ_pojisteni_id):
    typ_pojisteni = get_object_or_404(TypPojisteni, id = typ_pojisteni_id)
    typ_pojisteni.delete()
    messages.success(request, 'Typ pojištění byl úspěšně odstraněn.', extra_tags='deleted')
    return redirect('pridej-pojisteni')

@login_required
def pojisteni_detail(request, pojisteni_id):
    pojisteni = get_object_or_404(Pojisteni, id=pojisteni_id)
    return render(request, 'pojistenci/pojisteni_detail.html', {'pojisteni': pojisteni})

@login_required
def uprav_pojisteni(request, pojisteni_id):
    pojisteni = get_object_or_404(Pojisteni, id = pojisteni_id)
    if request.method == 'POST':
        form = PojisteniForm(request.POST, instance=pojisteni)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pojištění bylo úspěšně upraveno.')
            return redirect ('pojistenec_detail', id=pojisteni.pojistenec.id)
    else:
        form = PojisteniForm(instance=pojisteni)

    return render(request, 'pojistenci/uprav_pojisteni.html', {'form': form, 'pojisteni': pojisteni})

@login_required
def smaz_pojisteni(request, pojisteni_id):
    pojisteni = get_object_or_404(Pojisteni, id = pojisteni_id)
    pojisteni.delete()
    messages.success(request, 'Pojištění bylo úspěšně odstraněno.', extra_tags='deleted')
    return redirect ('pojistenec_detail', id=pojisteni.pojistenec.id)

@staff_member_required(login_url='login')
def typ_pojisteni(request):
    typy_pojisteni = TypPojisteni.objects.all()
    if request.method == 'POST':
        form = TypPojisteniForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Nový typ pojištění byl úspěšně přidán.')
            return redirect('pridej-typ-pojisteni')
    else:
        form = TypPojisteniForm()

    return render(request, 'pojistenci/typy_pojisteni.html', {'form': form,'typy_pojisteni': typy_pojisteni})

class VlastniLoginView(LoginView):
    template_name = 'pojistenci/login.html'
    form_class = VlastniLoginForm


def registrace(request):
    if request.method == 'POST':
        form = RegistraceForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)

            # ----- DEV REŽIM -----
            if settings.APP_ENV == "dev":
                user.is_active = True
                user.save()
                login(request, user)
                return redirect('homepage')

            # ----- PROD REŽIM -----
            user.is_active = False
            user.save()

            uidb64 = urlsafe_base64_encode(force_bytes(user.pk))
            token = default_token_generator.make_token(user)

            current_site = get_current_site(request)
            activation_link = f"http://{current_site.domain}{reverse('activate', args=[uidb64, token])}"

            subject = 'Aktivujte si účet'
            message = render_to_string(
                'registration/activation_email.txt',
                {'user': user, 'activation_link': activation_link}
            )

            email = EmailMessage(subject, message, settings.DEFAULT_FROM_EMAIL, to=[user.email])
            email.send(fail_silently=False)

            return render(request, 'pojistenci/registration_complete.html', {'form': form})

    else:
        form = RegistraceForm()

    return render(request, 'pojistenci/registrace.html', {'form': form})

def activate(request, uidb64, token):
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is not None and default_token_generator.check_token(user, token):
        user.is_active = True
        user.save()
        messages.success(request, 'Váš účet byl aktivován, nyní se můžete přihlásit.')
        return redirect('login')
    else:
        return render(request, 'pojistenci/activation_invalid.html')

@staff_member_required(login_url='login')
def bulk_upload_pojistenci(request):
    if request.method == 'POST':
        form = BulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_excel = request.FILES['file']
            try:
                wb = load_workbook(uploaded_excel, data_only=True)
                sheet = wb.active
            except Exception as e:
                form.add_error('file', f'Chyba při čtení Excelu: {e}')
            else:
                # první řádek = hlavička
                header = [cell.value for cell in sheet[1]]

                required_columns = ['Jméno','Příjmení','Ulice','Město','PSČ','Telefon','Email']
                missing = [col for col in required_columns if col not in header]

                if missing:
                    form.add_error('file', f'Excel postrádá sloupce: {", ".join(missing)}')
                else:
                    # mapa sloupců (název → index sloupce)
                    col_map = {name: header.index(name) for name in required_columns}

                    created = 0

                    try:
                        with transaction.atomic():
                            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                                
                                def normalize(value):
                                    if value is None:
                                        return ""
                                    return str(value).strip()

                                data = {
                                    'first_name':     normalize(row[col_map['Jméno']]),
                                    'last_name':      normalize(row[col_map['Příjmení']]),
                                    'address_street': normalize(row[col_map['Ulice']]),
                                    'address_city':   normalize(row[col_map['Město']]),
                                    'psc':            normalize(row[col_map['PSČ']]),
                                    'phone':          normalize(row[col_map['Telefon']]),
                                    'email':          normalize(row[col_map['Email']]),
                                }

                                obj = Pojistenec(**data)
                                obj.full_clean()
                                obj.save()
                                created += 1

                    except ValidationError as e:
                        form.add_error('file',
                            f"Chyba v řádku {row_index}: {e.messages[0]}"
                        )
                    else:
                        messages.success(request,
                            f"Úspěšně přidáno {created} pojištěnců."
                        )
                        return redirect('bulk-upload-pojistenci')
    else:
        form = BulkUploadForm()

    return render(request, 'pojistenci/bulk_upload.html', {'form': form})

@login_required
def vypis_api(request):
    return render(request, 'pojistenci/vypis_api.html')

def o_aplikaci(request):
    return render(request, 'pojistenci/o_aplikaci.html')

def vychytavky(request):
    return render(request, 'pojistenci/vychytavky.html')

@login_required
def generate_qr(request):
    if request.method == 'POST':
        data = request.POST.get('qr_text')
        qr_size = int(request.POST.get('qr_size'))
        qr_border = int(request.POST.get('qr_border'))
        qr_color = request.POST.get('qr_color')
        qr_background_color = request.POST.get('qr_background_color')
        use_custom_logo = request.POST.get("use_custom_logo") == "on"
        custom_logo_file = request.FILES.get("custom_logo_file")

        logo = None
        if use_custom_logo and custom_logo_file:
            try:
                logo = Image.open(custom_logo_file).convert("RGBA")
            except Exception as e:
                print("LOGO ERROR:", e)
                logo = None
        
        qr = qrcode.QRCode(version=None, box_size=qr_size, border=qr_border, error_correction=qrcode.constants.ERROR_CORRECT_H)
        qr.add_data(data)
        qr.make(fit=True)
        image = qr.make_image(fill_color=qr_color, back_color = qr_background_color).convert('RGBA')

        if logo:
            # spočítat cílovou velikost loga jako 20 % šířky QR
            qr_w, qr_h = image.size
            logo_target_width = int(qr_w * 0.20)

            wpercent = logo_target_width / float(logo.size[0])
            logo_target_height = int(float(logo.size[1]) * wpercent)

            logo = logo.resize((logo_target_width, logo_target_height), Image.LANCZOS)

            # pozice uprostřed
            pos = ((qr_w - logo_target_width) // 2, (qr_h - logo_target_height) // 2)

            image.alpha_composite(logo, dest=pos)
        
        # Uložení QR kódu do mediální složky
        filename = f"qr_code_{timezone.now().strftime('%Y%m%d%H%M%S')}.png"
        save_dir = os.path.join(settings.MEDIA_ROOT, "qr_codes")
        os.makedirs(save_dir, exist_ok=True)             # vytvoří složku pokud neexistuje
        full_path = os.path.join(save_dir, filename)
        image.save(full_path)

        qr_image_url = f"{settings.MEDIA_URL}qr_codes/{filename}"

        # Smazání starších QR kódů:
        try:
            files = sorted([os.path.join(save_dir, f) for f in os.listdir(save_dir)], key=os.path.getmtime)
            if len(files) > 10:  # ponechat posledních 10
                for old_file in files[:-10]:
                    os.remove(old_file)
        except Exception:
            pass  # pokud se něco pokazí, prostě to přeskočíme

        return render(request, 'pojistenci/generate_qr.html', {'qr_image_url': qr_image_url, 'data': data})
    
    return render(request, 'pojistenci/generate_qr.html')

@login_required
def generate_chart(request):
    DEFAULT_COLORS = ["#004289", "#5C9EAE", "#8F2D56", "#E9BA6A", "#E4E5C3","#3F7CAC", "#D9BF77", "#6B4226", "#A1C181", "#F2E394"]
    if request.method == "POST":
        count = int(request.POST.get("donut_number"))
        values = list(map(int, request.POST.getlist("chart_values")))
        
        if sum(values) != 100:
            return render(request,'pojistenci/generate_donut_chart.html',
                {
                    "error_message": f"Součet všech hodnot jednotlivých položek musí být 100. Aktuálně je to {sum(values)}.",
                    "prev_count": count,
                    "prev_values": values,
                    "prev_use_custom": request.POST.get("use_custom_colors") == "on",
                    "prev_colors": request.POST.getlist("chart_colors") or DEFAULT_COLORS[:count]
                }
            )

        use_custom = request.POST.get("use_custom_colors") == "on"
        prev_colors = request.POST.getlist("chart_colors")
        if use_custom:
            if not prev_colors or len(prev_colors) < count:
                prev_colors = DEFAULT_COLORS[:count]
            colors = prev_colors[:count]
        else:
            colors = DEFAULT_COLORS[:count]
        
        fig, ax = plt.subplots()
        ax.pie(values, colors=colors[:len(values)],wedgeprops=dict(width=0.5), startangle=-40)
        centre_circle = plt.Circle((0,0),0.25,fc='none')
        fig.gca().add_artist(centre_circle)
        ax.axis('equal')  
        # plt.title(f'Donut Chart with {count} segments')
        fig.patch.set_alpha(0)  # transparentní pozadí
        ax.patch.set_alpha(0)

        filename = f"donut_chart_{timezone.now().strftime('%Y%m%d%H%M%S')}.png"
        save_dir = os.path.join(settings.MEDIA_ROOT, "donut_charts")
        os.makedirs(save_dir, exist_ok=True)
        full_path = os.path.join(save_dir, filename)
        
        plt.savefig(full_path, transparent=True)
        plt.close()

        donut_image_url = f"{settings.MEDIA_URL}donut_charts/{filename}"

        # Smazání starších grafů:
        try:
            files = sorted([os.path.join(save_dir, f) for f in os.listdir(save_dir)], key=os.path.getmtime)
            if len(files) > 10:  # ponechat posledních 10
                for old_file in files[:-10]:
                    os.remove(old_file)
        except Exception:
            pass  # pokud se něco pokazí, prostě to přeskočíme

        return render(request, 'pojistenci/generate_donut_chart.html', {'donut_image_url': donut_image_url, 'prev_count': count, "prev_values": values, "prev_colors": colors, "prev_use_custom": use_custom})
    
    return render(request, 'pojistenci/generate_donut_chart.html', {"prev_count": 4, "prev_values": [25, 25, 25, 25], "prev_colors": DEFAULT_COLORS[:4], "prev_use_custom": False})


@login_required
def remove_background(request):
    if request.method == "POST":
        image_file = request.FILES.get("image_file")
        if not image_file:
            return render(request, "pojistenci/remove_background.html", {
                "error_message": "Prosím nahraj fotografii."
            })

        api_key = os.getenv("REMOVE_BG_API_KEY")
        if not api_key:
            return render(request, "pojistenci/remove_background.html", {
                "error_message": "Chybí API klíč k remove.bg. Zkontroluj .env."
            })

        # ==== CALL REMOVE.BG API ====
        files = {"image_file": image_file}
        data = {"size": "auto"}   # může být "auto", "medium", "full"

        response = requests.post(
            "https://api.remove.bg/v1.0/removebg",
            data=data,
            files=files,
            headers={"X-Api-Key": api_key},
        )

        if response.status_code != 200:
            return render(request, "pojistenci/remove_background.html", {
                "error_message": f"Remove.bg API error: {response.text}"
            })

        # ==== SAVE RESULT ====
        output_dir = os.path.join(settings.MEDIA_ROOT, "removed_backgrounds")
        os.makedirs(output_dir, exist_ok=True)

        filename = f"no_bg_{timezone.now().strftime('%Y%m%d%H%M%S')}.png"
        full_path = os.path.join(output_dir, filename)

        with open(full_path, "wb") as f:
            f.write(response.content)

        # === zde přidáme rotaci starších obrázků ===
        try:
            files = sorted(
                [os.path.join(output_dir, f) for f in os.listdir(output_dir)],
                key=os.path.getmtime
            )
            if len(files) > 5:
                for old_file in files[:-5]:
                    os.remove(old_file)
        except Exception:
            pass  # pokud se něco pokazí, prostě to přeskočíme

        output_url = f"{settings.MEDIA_URL}removed_backgrounds/{filename}"

        return render(request, "pojistenci/remove_background.html", {
            "output_image_url": output_url
        })

    return render(request, "pojistenci/remove_background.html")

@login_required
def generate_pie_chart(request):
    DEFAULT_COLORS = ["#004289", "#5C9EAE", "#8F2D56", "#E9BA6A", "#E4E5C3","#3F7CAC", "#D9BF77", "#6B4226", "#A1C181", "#F2E394"    ]
    if request.method == "POST":
        count = int(request.POST.get("pie-chart-number"))
        values = list(map(int, request.POST.getlist("chart_values")))
        highlight_index = None
        highlight_outline = False
        full_outline = False

        if request.user.is_staff:
            # index vystouplé části, může být None
            raw_index = request.POST.get("highlight_index")
            if raw_index not in (None, "", "none"):
                highlight_index = int(raw_index)

        full_outline = request.POST.get("full_outline") == "on"
        
        if sum(values) != 100:
            return render(request,'pojistenci/generate_pie_chart.html',
                {
                    "error_message": f"Součet všech hodnot jednotlivých položek musí být 100. Aktuálně je to {sum(values)}.",
                    "prev_count": count,
                    "prev_values": values,
                    "prev_use_custom": request.POST.get("use_custom_colors") == "on",
                    "prev_colors": request.POST.getlist("chart_colors") or DEFAULT_COLORS[:count]
                }
            )

        use_custom = request.POST.get("use_custom_colors") == "on"
        prev_colors = request.POST.getlist("chart_colors")
        if use_custom:
            if not prev_colors or len(prev_colors) < count:
                prev_colors = DEFAULT_COLORS[:count]
            colors = prev_colors[:count]
        else:
            colors = DEFAULT_COLORS[:count]

        labels = request.POST.getlist("chart_labels")
        use_labels = request.POST.get("use_labels") == "on"
        if use_labels:
            if len(labels) < count:
                labels = labels + [""] * (count - len(labels))
        else:
            labels = [""] * count


        explode = [0] * len(values)
        if highlight_index is not None:
            explode[highlight_index] = 0.12
        
        fig, ax = plt.subplots(figsize=(7,7))
        wedges, texts = ax.pie(
            values,
            explode=explode,
            colors=colors[:len(values)],
            labels=labels,
            startangle=-40,
            textprops={'fontsize': 14, 'color': '#000'}
        )

        if full_outline:
            # Nakresli kruh mírně větší než pie graf
            circ = plt.Circle((0, 0), radius=1.08, fill=False, linewidth=9, edgecolor="#55002e")
            ax.add_patch(circ)

        # Posunout popisky dál od středu, aby nebyly uříznuté
        for t in texts:
            x, y = t.get_position()
            t.set_position((x * 1.15, y * 1.15))

        ax.axis('equal') 
        plt.subplots_adjust(left=0.05, right=0.95, top=0.95, bottom=0.05)
        fig.patch.set_alpha(0)  # transparentní pozadí
        ax.patch.set_alpha(0)

        filename = f"pie_chart_{timezone.now().strftime('%Y%m%d%H%M%S')}.png"
        save_dir = os.path.join(settings.MEDIA_ROOT, "pie_charts")
        os.makedirs(save_dir, exist_ok=True)
        full_path = os.path.join(save_dir, filename)
        
        plt.savefig(full_path, transparent=True, dpi=150)
        plt.close()

        pie_chart_image_url = f"{settings.MEDIA_URL}pie_charts/{filename}"

        # Smazání starších grafů:
        try:
            files = sorted([os.path.join(save_dir, f) for f in os.listdir(save_dir)], key=os.path.getmtime)
            if len(files) > 10:  # ponechat posledních 10
                for old_file in files[:-10]:
                    os.remove(old_file)
        except Exception:
            pass  # pokud se něco pokazí, prostě to přeskočíme

        return render(request, 'pojistenci/generate_pie_chart.html', {'pie_chart_image_url': pie_chart_image_url, 'prev_count': count, "prev_values": values, "prev_colors": colors, "prev_use_custom": use_custom, "prev_labels": labels, "prev_use_labels": use_labels, "is_staff": request.user.is_staff, "count_range": range(count)})
    return render(request, 'pojistenci/generate_pie_chart.html', {"is_staff": request.user.is_staff, "count_range": range(3), "prev_count" : 3})

@login_required
def generate_password(request):
    generated_passwords = []
    error_message = None

    if request.method == "POST":
        try:
            length = int(request.POST.get("password_length", 16))
            if length < 1 or length > 100:
                raise ValueError
        except ValueError:
            error_message = "Délka hesla musí být číslo v rozsahu 1–100."
        else:
            use_upper = request.POST.get("use_uppercase") == "on"
            use_lower = request.POST.get("use_lowercase") == "on"
            use_digits = request.POST.get("use_digits") == "on"
            use_special = request.POST.get("use_special") == "on"

            import string
            import secrets

            character_pool = ""

            if use_upper:
                character_pool += string.ascii_uppercase
            if use_lower:
                character_pool += string.ascii_lowercase
            if use_digits:
                character_pool += string.digits
            if use_special:
                character_pool += "!@#$%^&*()-_=+[]{}?"

            if not character_pool:
                error_message = "Vyberte alespoň jednu kategorii znaků."
            else:
                for _ in range(5):
                    generated_passwords.append(
                        "".join(secrets.choice(character_pool) for _ in range(length))
                    )

    return render(
        request,
        "pojistenci/generate_password.html",
        {
            "generated_passwords": generated_passwords,
            "error_message": error_message,
        },
    )


@login_required
def generate_grouped_bar(request):
    template = "pojistenci/generate_grouped_bar.html"
    ctx = {}

    if request.method == "POST":
        try:
            # počty
            groups_count = int(request.POST.get("grouped-bar-groups-number", 5))
            items_count  = int(request.POST.get("grouped-bar-items-number", 5))
            groups_count = max(1, min(5, groups_count))
            items_count  = max(1, min(10, items_count))

            # osa Y
            y_min  = float(str(request.POST.get("y-min", "0")).replace(",", "."))
            y_max  = float(str(request.POST.get("y-max", "100")).replace(",", "."))
            y_step = float(str(request.POST.get("y-step", "10")).replace(",", "."))
            hide_legend = request.POST.get("hide-legend") == "on"
            hide_group_labels = request.POST.get("hide-group-labels") == "on"

            if y_max <= y_min:
                raise ValueError("Max Y musí být větší než Min Y.")
            if y_step <= 0:
                raise ValueError("Krok osy Y musí být kladné číslo.")

            # labels skupin
            group_labels = []
            for g in range(groups_count):
                val = (request.POST.get(f"group_label_{g}", "") or "").strip()
                group_labels.append(val if val else f"Skupina {g+1}")

            # labels + barvy položek
            item_labels = []
            item_colors = []
            for i in range(items_count):
                il = (request.POST.get(f"item_label_{i}", "") or "").strip()
                ic = (request.POST.get(f"item_color_{i}", "#00469B") or "").strip()
                item_labels.append(il if il else f"Položka {i+1}")
                item_colors.append(ic if ic else "#00469B")

            # hodnoty
            values = []
            for g in range(groups_count):
                row = []
                for i in range(items_count):
                    raw = (request.POST.get(f"value_{g}_{i}", "") or "").strip()
                    if raw == "":
                        raise ValueError(f"Chybí hodnota pro {group_labels[g]} / {item_labels[i]}.")
                    v = float(raw.replace(",", "."))
                    if v < y_min or v > y_max:
                        raise ValueError(
                            f"Hodnota {v} je mimo rozsah {y_min}–{y_max} ({group_labels[g]} / {item_labels[i]})."
                        )
                    row.append(v)
                values.append(row)

            # --- kreslení ---
            fig, ax = plt.subplots(figsize=(10, 5), dpi=150)
            for spine in ax.spines.values():
                spine.set_visible(False)

            group_positions = list(range(groups_count))
            total_width = 0.65
            inner_gap = 0.02
            bar_width = (total_width - inner_gap * (items_count - 1)) / items_count

            for i in range(items_count):
                x = [gp - total_width/2 + bar_width/2 + i * (bar_width + inner_gap) for gp in group_positions]
                y = [values[g][i] for g in range(groups_count)]
                ax.bar(x, y, width=bar_width, label=item_labels[i], color=item_colors[i])

            ax.set_xticks(group_positions)

            if hide_group_labels:
                ax.set_xticklabels([])              # žádné názvy
                ax.tick_params(axis="x", length=0)  # zruší i “čárky” na ose
            else:
                ax.set_xticklabels(group_labels)

            # y ticks
            ticks = []
            t = y_min
            while t <= y_max + 1e-9:
                ticks.append(t)
                t += y_step
            ax.set_yticks(ticks)

            ax.tick_params(axis="y", length=0)

            ax.set_ylabel("")
            ax.grid(axis="y", linestyle="--", alpha=0.35)
            if not hide_legend:
                ax.legend(ncols=min(items_count, 5), fontsize=9)
            fig.tight_layout()

            # --- uložení ---
            filename = f"grouped_bar_{timezone.now().strftime('%Y%m%d%H%M%S')}.png"
            save_dir = os.path.join(settings.MEDIA_ROOT, "grouped_bars")
            os.makedirs(save_dir, exist_ok=True)
            full_path = os.path.join(save_dir, filename)

            fig.savefig(full_path, format="png", transparent=True)
            plt.close(fig)

            grouped_bar_image_url = f"{settings.MEDIA_URL}grouped_bars/{filename}"
            ctx["chart_url"] = grouped_bar_image_url

            # cleanup (necháme posledních 10)
            try:
                files = sorted(
                    [os.path.join(save_dir, f) for f in os.listdir(save_dir)],
                    key=os.path.getmtime
                )
                if len(files) > 10:
                    for old_file in files[:-10]:
                        os.remove(old_file)
            except Exception:
                pass

            # prefills zpět do šablony
            ctx["prev_groups_count"] = groups_count
            ctx["prev_items_count"] = items_count
            ctx["prev_y_min"] = str(y_min).replace(",", ".")
            ctx["prev_y_max"] = str(y_max).replace(",", ".")
            ctx["prev_y_step"] = str(y_step).replace(",", ".")
            ctx["prev_group_labels"] = group_labels
            ctx["prev_item_labels"] = item_labels
            ctx["prev_item_colors"] = item_colors
            ctx["prev_values"] = values
            ctx["prev_hide_legend"] = hide_legend
            ctx["prev_hide_group_labels"] = hide_group_labels

        except Exception as e:
            ctx["error_message"] = str(e)
            ctx["prev_groups_count"] = request.POST.get("grouped-bar-groups-number", 2)
            ctx["prev_items_count"] = request.POST.get("grouped-bar-items-number", 3)

            ctx["prev_y_min"] = str(request.POST.get("y-min", "0")).replace(",", ".")
            ctx["prev_y_max"] = str(request.POST.get("y-max", "100")).replace(",", ".")
            ctx["prev_y_step"] = str(request.POST.get("y-step", "10")).replace(",", ".")

            ctx["prev_hide_legend"] = request.POST.get("hide-legend") == "on"
            ctx["prev_hide_group_labels"] = request.POST.get("hide-group-labels") == "on"

    return render(request, template, ctx)

@login_required
def generate_bar_chart(request):
    template = "pojistenci/generate_bar_chart.html"
    ctx = {}

    if request.method == "POST":
        try:
            # počty
            bars_count = int(request.POST.get("bar-chart-bars-number", 5))
            bars_count = max(1, min(15, bars_count))

            # osa Y
            y_min = float(str(request.POST.get("y-min", "0")).replace(",", "."))
            y_max = float(str(request.POST.get("y-max", "100")).replace(",", "."))
            y_step = float(str(request.POST.get("y-step", "10")).replace(",", "."))
            hide_bar_labels = request.POST.get("hide-bar-labels") == "on"

            if y_max <= y_min:
                raise ValueError("Max Y musí být větší než Min Y.")
            if y_step <= 0:
                raise ValueError("Krok osy Y musí být kladné číslo.")

            # labely a barvy
            bar_labels = []
            bar_colors = []

            for i in range(bars_count):
                label = (request.POST.get(f"bar_label_{i}", "") or "").strip()
                bar_labels.append(label if label else f"Sloupec {i+1}")

                col = (request.POST.get(f"bar_color_{i}") or "#00469B").strip()
                bar_colors.append(col)

            # hodnoty
            values = []
            for i in range(bars_count):
                raw = (request.POST.get(f"bar_value_{i}", "") or "").strip()
                if raw == "":
                    raise ValueError(f"Chybí hodnota pro {bar_labels[i]}.")
                v = float(raw.replace(",", "."))
                if v < y_min or v > y_max:
                    raise ValueError(
                        f"Hodnota {v} je mimo rozsah {y_min}–{y_max} ({bar_labels[i]})."
                    )
                values.append(v)

            # --- kreslení ---
            fig, ax = plt.subplots(figsize=(10, 5), dpi=150)
            for spine in ax.spines.values():
                spine.set_visible(False)

            x_positions = list(range(bars_count))
            ax.bar(x_positions, values, color=bar_colors, width=0.6, zorder=3)

            ax.set_ylim(y_min, y_max)

            ax.set_xticks(x_positions)

            if hide_bar_labels:
                ax.set_xticklabels([])
                ax.tick_params(axis="x", length=0)
            else:
                ax.set_xticklabels(bar_labels, rotation=45, ha="right", rotation_mode="anchor")
                ax.tick_params(axis="x", length=0, labelsize=9)


            # y ticks
            ax.tick_params(axis="y", length=0)
            ticks = []
            t = y_min
            while t <= y_max + 1e-9:
                ticks.append(t)
                t += y_step
            ax.set_yticks(ticks)

            ax.set_ylabel("")
            ax.grid(axis="y", linestyle="-", alpha=0.35)
            ax.set_axisbelow(True)   # <-- grid pod sloupce
            fig.tight_layout()

            # --- uložení ---
            filename = f"bar_chart_{timezone.now().strftime('%Y%m%d%H%M%S')}.png"
            save_dir = os.path.join(settings.MEDIA_ROOT, "bar_charts")
            os.makedirs(save_dir, exist_ok=True)
            full_path = os.path.join(save_dir, filename)

            fig.savefig(full_path, format="png", transparent=True)
            plt.close(fig)

            bar_chart_image_url = f"{settings.MEDIA_URL}bar_charts/{filename}"
            ctx["chart_url"] = bar_chart_image_url

            # cleanup (necháme posledních 10)
            try:
                files = sorted(
                    [os.path.join(save_dir, f) for f in os.listdir(save_dir)],
                    key=os.path.getmtime
                )
                if len(files) > 10:
                    for old_file in files[:-10]:
                        os.remove(old_file)
            except Exception:
                pass

            # prefills zpět do šablony
            ctx["prev_bars_count"] = bars_count
            ctx["prev_y_min"] = str(y_min).replace(",", ".")
            ctx["prev_y_max"] = str(y_max).replace(",", ".")
            ctx["prev_y_step"] = str(y_step).replace(",", ".")
            ctx["prev_bar_labels"] = bar_labels
            # ctx["prev_bar_color"] = request.POST.get("bar_color", "#3b82f6")
            ctx["prev_values"] = values
            ctx["prev_hide_bar_labels"] = hide_bar_labels
            ctx["prev_bar_colors"] = bar_colors

        except Exception as e:
            ctx["error_message"] = str(e)
            ctx["prev_bars_count"] = request.POST.get("bar-chart-bars-number", 5)
            ctx["prev_y_min"] = str(request.POST.get("y-min", "0")).replace(",", ".")
            ctx["prev_y_max"] = str(request.POST.get("y-max", "100")).replace(",", ".")
            ctx["prev_y_step"] = str(request.POST.get("y-step", "10")).replace(",", ".")
            ctx["prev_hide_bar_labels"] = request.POST.get("hide-bar-labels") == "on"
            ctx["prev_bar_colors"] = []

    return render(request, template, ctx)

@login_required
def convert_csv_to_xlsx(request):
    template = "pojistenci/convert_csv_to_xlsx.html"
    ctx = {}

    if request.method == "POST":
        try:
            up = request.FILES.get("csv_file")
            if not up:
                raise ValueError("Nahrajte prosím CSV soubor.")

            # jednoduchá validace typu/názvu
            filename = (up.name or "").lower()
            if not filename.endswith(".csv"):
                raise ValueError("Soubor musí být ve formátu .csv")

            # pojistka velikosti (klidně uprav)
            max_mb = 5
            if up.size and up.size > max_mb * 1024 * 1024:
                raise ValueError(f"Soubor je příliš velký (max {max_mb} MB).")

            # --- konverze ---
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"

            # TextIOWrapper: přečteme upload jako text
            # encoding: utf-8-sig zvládne i BOM z Excelu
            wrapper = TextIOWrapper(up.file, encoding="utf-8-sig", newline="")

            # delimiter: buď zkusíme sniff, nebo fallback na ;
            sample = wrapper.read(4096)
            wrapper.seek(0)

            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
            except Exception:
                dialect = csv.excel
                dialect.delimiter = ";"

            reader = csv.reader(wrapper, dialect=dialect)

            for row in reader:
                # row je list stringů; openpyxl to vezme
                ws.append(row)

            # --- uložení ---
            ts = timezone.now().strftime("%Y%m%d%H%M%S%f")
            out_name = f"csv_to_xlsx_{ts}.xlsx"

            save_dir = os.path.join(settings.MEDIA_ROOT, "csv_to_xlsx")
            os.makedirs(save_dir, exist_ok=True)

            full_path = os.path.join(save_dir, out_name)
            wb.save(full_path)

            ctx["result_url"] = f"{settings.MEDIA_URL}csv_to_xlsx/{out_name}"
            ctx["result_filename"] = out_name

            # cleanup (necháme posledních 5)
            try:
                files = sorted(
                    [os.path.join(save_dir, f) for f in os.listdir(save_dir) if f.lower().endswith(".xlsx")],
                    key=os.path.getmtime
                )
                if len(files) > 5:
                    for old_file in files[:-5]:
                        os.remove(old_file)
            except Exception:
                pass

        except Exception as e:
            ctx["error_message"] = str(e)

    return render(request, template, ctx)


# tady budou funkce spojené s rozesílačem emailů:

@staff_member_required
def rozesilac_dashboard(request):
    cards = [
        {"title": "Šablony", "desc": "Vytvořit a upravit HTML šablony emailů.", "url": "rozesilac_templates"},
        {"title": "Kontakty", "desc": "Správa kontaktů + import z XLSX.", "url": "rozesilac_contacts"},
        {"title": "Odeslat", "desc": "Vybrat šablonu, příjemce a odeslat.", "url": "rozesilac_send"},
        {"title": "Kampaně", "desc": "Historie rozesílek, výsledky a chyby.", "url": "rozesilac_campaigns"},
    ]
    return render(request, "pojistenci/rozesilac/dashboard.html", {"cards": cards})


@staff_member_required
def rozesilac_templates(request):
    templates = EmailTemplate.objects.order_by("-updated_at", "name")
    return render(request, "pojistenci/rozesilac/templates_list.html", {"templates": templates})


@staff_member_required
def rozesilac_template_create(request):
    if request.method == "POST":
        form = EmailTemplateForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Šablona byla vytvořena.")
            return redirect("rozesilac_templates")
    else:
        form = EmailTemplateForm()

    recent_images = EmailImage.objects.all()[:4]
    total_size = EmailImage.objects.aggregate(total=Sum("file_size"))["total"] or 0
    limit_size = 100 * 1024 * 1024
    image_upload_form = EmailImageUploadForm()

    return render(
        request,
        "pojistenci/rozesilac/template_form.html",
        {
            "form": form,
            "page_title": "Nová šablona",
            "submit_label": "Vytvořit šablonu",
            "recent_images": recent_images,
            "total_size": total_size,
            "limit_size": limit_size,
            "image_upload_form": image_upload_form,
        },
    )

@staff_member_required
def rozesilac_template_edit(request, template_id):
    template_obj = get_object_or_404(EmailTemplate, id=template_id)

    if request.method == "POST":
        form = EmailTemplateForm(request.POST, instance=template_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Šablona byla upravena.")
            return redirect("rozesilac_templates")
    else:
        form = EmailTemplateForm(instance=template_obj)

    recent_images = EmailImage.objects.all()[:4]
    total_size = EmailImage.objects.aggregate(total=Sum("file_size"))["total"] or 0
    limit_size = 100 * 1024 * 1024
    image_upload_form = EmailImageUploadForm()

    return render(
        request,
        "pojistenci/rozesilac/template_form.html",
        {
            "form": form,
            "page_title": f"Upravit šablonu: {template_obj.name}",
            "submit_label": "Uložit změny",
            "template_obj": template_obj,
            "recent_images": recent_images,
            "total_size": total_size,
            "limit_size": limit_size,
            "image_upload_form": image_upload_form,
        },
    )


@staff_member_required
def rozesilac_template_delete(request, template_id):
    template_obj = get_object_or_404(EmailTemplate, id=template_id)

    if request.method == "POST":
        template_obj.delete()
        messages.success(request, "Šablona byla smazána.")
        return redirect("rozesilac_templates")

    return render(
        request,
        "pojistenci/rozesilac/template_delete.html",
        {"template_obj": template_obj},
    )

#pomocná funkce
def get_contact_salutation(contact):
    if contact.salutation and contact.salutation.strip():
        return contact.salutation.strip()
    if contact.name and contact.name.strip():
        return contact.name.strip()
    return contact.email

@staff_member_required
def rozesilac_contacts(request):
    if request.method == "POST" and request.POST.get("action") == "delete_contact":
        contact_id = request.POST.get("contact_id")
        Contact.objects.filter(id=contact_id).delete()
        messages.success(request, "Kontakt smazán.")
        return redirect("rozesilac_contacts")

    if request.method == "POST" and request.POST.get("action") == "add_group":
        group_form = ContactGroupForm(request.POST)
        if group_form.is_valid():
            group_form.save()
            messages.success(request, "Skupina byla vytvořena.")
            return redirect("rozesilac_contacts")
    else:
        group_form = ContactGroupForm()

    if request.method == "POST" and request.POST.get("action") == "delete_group":
        group_id = request.POST.get("group_id")
        ContactGroup.objects.filter(id=group_id).delete()
        messages.success(request, "Skupina byla smazána.")
        return redirect("rozesilac_contacts")

    add_form = ContactForm()
    import_form = ContactImportForm()

    if request.method == "POST" and request.POST.get("action") == "add_contact":
        add_form = ContactForm(request.POST)
        if add_form.is_valid():
            try:
                add_form.save()
                messages.success(request, "Kontakt uložen.")
                return redirect("rozesilac_contacts")
            except IntegrityError:
                add_form.add_error("email", "Tento email už v kontaktech existuje.")

    if request.method == "POST" and request.POST.get("action") == "import":
        import_form = ContactImportForm(request.POST, request.FILES)
        if import_form.is_valid():
            f = import_form.cleaned_data["file"]
            selected_group = import_form.cleaned_data.get("group")

            wb = load_workbook(filename=f, data_only=True)
            ws = wb.active

            header_row = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]

            def find_col(possible_names):
                for i, h in enumerate(header_row):
                    if h in possible_names:
                        return i
                return None

            name_col = find_col({"jméno", "jmeno", "name"})
            email_col = find_col({"email", "e-mail", "e mail", "mail"})
            salutation_col = find_col({"osloveni", "salutation", "pozdrav", "oslovení"})

            if name_col is None or email_col is None:
                messages.error(request, "XLSX musí mít v prvním řádku sloupce 'jméno' a 'email'.")
                return redirect("rozesilac_contacts")

            created = 0
            skipped = 0
            invalid = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                raw_email = row[email_col] if email_col < len(row) else None
                raw_name = row[name_col] if name_col < len(row) else None
                raw_salutation = row[salutation_col] if (salutation_col is not None and salutation_col < len(row)) else None

                email = (str(raw_email).strip() if raw_email is not None else "").lower()
                name = str(raw_name).strip() if raw_name is not None else ""
                salutation = str(raw_salutation).strip() if raw_salutation is not None else ""

                if not email:
                    continue

                try:
                    validate_email(email)
                except ValidationError:
                    invalid += 1
                    continue

                obj, was_created = Contact.objects.get_or_create(
                    email=email,
                    defaults={"name": name, "salutation": salutation, "is_active": True},
                )

                if was_created:
                    created += 1
                else:
                    skipped += 1

                if selected_group:
                    obj.groups.add(selected_group)

            messages.success(
                request,
                f"Import hotový. Přidáno: {created}, přeskočeno (duplicitní): {skipped}, neplatné emaily: {invalid}."
            )
            return redirect("rozesilac_contacts")

    contacts = Contact.objects.prefetch_related("groups").order_by("groups__name", "email").distinct()
    groups = ContactGroup.objects.all()

    return render(
        request,
        "pojistenci/rozesilac/contacts_list.html",
        {
            "contacts": contacts,
            "groups": groups,
            "add_form": add_form,
            "import_form": import_form,
            "group_form": group_form,
        },
    )

@staff_member_required
def rozesilac_contact_edit(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)

    if request.method == "POST":
        form = ContactForm(request.POST, instance=contact)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Kontakt byl upraven.")
                return redirect("rozesilac_contacts")
            except IntegrityError:
                form.add_error("email", "Tento email už v kontaktech existuje.")
    else:
        form = ContactForm(instance=contact)

    return render(
        request,
        "pojistenci/rozesilac/contact_form.html",
        {
            "form": form,
            "page_title": f"Upravit kontakt: {contact.email}",
            "submit_label": "Uložit změny",
            "contact_obj": contact,
        },
    )



@staff_member_required
def rozesilac_send(request):

    if request.method == "POST":
        form = SendCampaignForm(request.POST)

        if form.is_valid():

            template = form.cleaned_data["template"]
            send_mode = form.cleaned_data["send_mode"]
            test_email = form.cleaned_data.get("test_email")
            contacts = form.cleaned_data.get("contacts")
            note = form.cleaned_data.get("note", "")
            from_email = form.cleaned_data.get("from_email") or settings.NEWSLETTER_DEFAULT_FROM_EMAIL

            is_test = send_mode == "test"

            campaign = EmailCampaign.objects.create(
                template=template,
                created_by=request.user,
                subject=template.subject,
                html_body=template.html_body,
                text_body=template.text_body,
                is_test=is_test,
                note=note,
            )

            # --------------------------------------------------
            # připravíme seznam příjemců
            # --------------------------------------------------

            recipients = []

            if is_test:
                recipients.append({
                    "email": test_email,
                    "name": "",
                    "contact": None,
                })
            else:
                for contact in contacts:
                    recipients.append({
                        "email": contact.email,
                        "name": contact.name,
                        "contact": contact,
                    })

            sent_count = 0
            failed_count = 0

            # základ domény pro generování absolutních URL
            base_url = f"{request.scheme}://{request.get_host()}"

            # --------------------------------------------------
            # odesílání
            # --------------------------------------------------

            for recipient in recipients:

                delivery = EmailDelivery.objects.create(
                    campaign=campaign,
                    to_email=recipient["email"],
                    to_name=recipient["name"],
                    status="queued",
                )

                try:
                    contact = recipient.get("contact")
                    osloveni = get_contact_salutation(contact) if contact else recipient["email"]

                    if contact:
                        unsubscribe_path = reverse("rozesilac_unsubscribe", args=[contact.unsubscribe_token],)
                        unsubscribe_url = f"{base_url}{unsubscribe_path}"
                    else:
                        unsubscribe_url = ""

                    template_context = Context({
                        "osloveni": osloveni,
                        "jmeno": contact.name if contact and contact.name else "",
                        "email": recipient["email"],
                        "unsubscribe_url": unsubscribe_url,
                    })

                    rendered_subject = Template(campaign.subject).render(template_context)
                    rendered_html_body = Template(campaign.html_body).render(template_context)

                    text_template = campaign.text_body.strip() if campaign.text_body else ""
                    if text_template:
                        rendered_text_body = Template(text_template).render(template_context)
                    else:
                        rendered_text_body = "Tento email obsahuje HTML verzi zprávy."

                    # --------------------------------------------------
                    # DEV = klasický Django email backend
                    # --------------------------------------------------

                    if settings.APP_ENV != "prod":

                        msg = EmailMultiAlternatives(
                            subject=rendered_subject,
                            body=rendered_text_body,
                            from_email=from_email,
                            to=[recipient["email"]],
                            reply_to=["info@liedersociety.cz"],
                        )

                        msg.attach_alternative(rendered_html_body, "text/html")
                        msg.send(fail_silently=False)

                    # --------------------------------------------------
                    # PROD = Brevo API
                    # --------------------------------------------------

                    else:

                        payload = {
                            "sender": {
                                "email": from_email,
                                "name": "Lieder Society",
                            },
                            "to": [
                                {
                                    "email": recipient["email"],
                                    "name": recipient["name"] or "",
                                }
                            ],
                            "subject": rendered_subject,
                            "htmlContent": rendered_html_body,
                            "textContent": rendered_text_body,
                            "replyTo": {
                                "email": "info@liedersociety.cz",
                                "name": "Lieder Society",
                            },
                        }

                        headers = {
                            "accept": "application/json",
                            "api-key": settings.BREVO_API_KEY,
                            "content-type": "application/json",
                        }

                        response = requests.post(
                            settings.BREVO_API_URL,
                            json=payload,
                            headers=headers,
                            timeout=20,
                        )

                        if response.status_code >= 400:
                            raise Exception(f"Brevo error {response.status_code}: {response.text}")

                    delivery.status = "sent"
                    delivery.sent_at = timezone.now()
                    delivery.error = ""
                    delivery.save(update_fields=["status", "sent_at", "error"])

                    sent_count += 1

                except Exception as exc:

                    delivery.status = "failed"
                    delivery.error = str(exc)
                    delivery.save(update_fields=["status", "error"])

                    failed_count += 1

            # --------------------------------------------------
            # zpráva pro uživatele
            # --------------------------------------------------

            if failed_count == 0:
                messages.success(
                    request,
                    f"Odeslání dokončeno. Úspěšně odesláno: {sent_count}."
                )
            else:
                messages.warning(
                    request,
                    f"Odeslání dokončeno s chybami. Odesláno: {sent_count}, chyb: {failed_count}."
                )

            return redirect("rozesilac_campaign_detail", campaign_id=campaign.id)

    else:
        form = SendCampaignForm()

    return render(
        request,
        "pojistenci/rozesilac/send.html",
        {
            "form": form,
            "templates_for_preview": EmailTemplate.objects.all().order_by("name"),
        },
    )


@staff_member_required
def rozesilac_campaign_detail(request, campaign_id):
    campaign = get_object_or_404(EmailCampaign, id=campaign_id)
    deliveries = campaign.deliveries.all().order_by("created_at")

    sent_count = deliveries.filter(status="sent").count()
    failed_count = deliveries.filter(status="failed").count()
    queued_count = deliveries.filter(status="queued").count()

    return render(
        request,
        "pojistenci/rozesilac/campaign_detail.html",
        {
            "campaign": campaign,
            "deliveries": deliveries,
            "sent_count": sent_count,
            "failed_count": failed_count,
            "queued_count": queued_count,
        },
    )

@staff_member_required
def rozesilac_campaigns(request):
    campaigns = (
        EmailCampaign.objects
        .select_related("created_by", "template")
        .prefetch_related("deliveries")
        .order_by("-created_at")
    )

    campaign_rows = []
    for campaign in campaigns:
        deliveries = campaign.deliveries.all()
        total_count = deliveries.count()
        sent_count = deliveries.filter(status="sent").count()
        failed_count = deliveries.filter(status="failed").count()
        queued_count = deliveries.filter(status="queued").count()

        campaign_rows.append({
            "campaign": campaign,
            "total_count": total_count,
            "sent_count": sent_count,
            "failed_count": failed_count,
            "queued_count": queued_count,
        })

    return render(
        request,
        "pojistenci/rozesilac/campaigns_list.html",
        {"campaign_rows": campaign_rows},
    )

# rozesílač - nahrávání obrázků na server pro použití v šablonách:
@staff_member_required
def rozesilac_images(request):
    upload_form = EmailImageUploadForm()

    if request.method == "POST":
        if request.POST.get("action") == "upload":
            upload_form = EmailImageUploadForm(request.POST, request.FILES)

            if upload_form.is_valid():
                obj = upload_form.save(commit=False)
                obj.uploaded_by = request.user
                obj.file_size = obj.image.size
                obj.save()

                messages.success(request, "Obrázek byl nahrán.")
                return redirect("rozesilac_images")

        elif request.POST.get("action") == "delete":
            image_id = request.POST.get("image_id")
            obj = get_object_or_404(EmailImage, id=image_id)

            if obj.image:
                obj.image.delete(save=False)
            obj.delete()

            messages.success(request, "Obrázek byl smazán.")
            return redirect("rozesilac_images")

    images = EmailImage.objects.all()
    total_size = EmailImage.objects.aggregate(total=Sum("file_size"))["total"] or 0
    limit_size = 100 * 1024 * 1024

    return render(
        request,
        "pojistenci/rozesilac/images_gallery.html",
        {
            "upload_form": upload_form,
            "images": images,
            "total_size": total_size,
            "limit_size": limit_size,
        },
    )

@staff_member_required
def rozesilac_image_upload(request):
    if request.method != "POST":
        return HttpResponseForbidden("Pouze POST.")

    form = EmailImageUploadForm(request.POST, request.FILES)
    next_url = request.POST.get("next") or "rozesilac_templates"

    if form.is_valid():
        obj = form.save(commit=False)
        obj.uploaded_by = request.user
        obj.file_size = obj.image.size
        obj.save()
        messages.success(request, "Obrázek byl nahrán.")
    else:
        for error in form.non_field_errors():
            messages.error(request, error)
            
        for field_name, errors in form.errors.items():
            if field_name == "__all__":
                continue
            for error in errors:
                messages.error(request, error)

    return redirect(next_url)


def unsubscribe_view(request, token):
    contact = get_object_or_404(Contact, unsubscribe_token=token)
    if request.method == "POST":
        if contact.is_active:
            contact.is_active = False
            contact.save()
            send_mail(
                subject="Odhlášení z odběru newsletteru",
                message=f"Tohle je automatická zpráva z Vaňkova super rozesílače. Chci oznámit, že kontakt {contact.email} se odhlásil/a z odběru Lieder newsletteru. V rozesílači bude teď tento kontakt označen jako neaktivní (ale z kontaktů se nesmazal).",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=["newsletter@liedersociety.cz"],
                fail_silently=True,
            )
        return render(request, "pojistenci/rozesilac/unsubscribe_done.html", {"contact": contact})
    return render(request, "pojistenci/rozesilac/unsubscribe_confirm.html", {"contact": contact})
